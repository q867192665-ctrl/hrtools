#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
数据汇总与更新模块
功能：数据汇总、批量导入、数据导出、统计报表
"""

import sqlite3
import os
import csv
import json
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple

DATABASE_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'salary_system.db')


class DataManager:
    """数据管理器"""
    
    def __init__(self, db_path=None):
        self.db_path = db_path or DATABASE_PATH
    
    def get_db_connection(self):
        """获取数据库连接"""
        conn = sqlite3.connect(self.db_path, timeout=30)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout=30000")
        return conn
    
    def sync_all_to_summary(self) -> Dict:
        conn = None
        try:
            conn = self.get_db_connection()
            cursor = conn.cursor()

            cursor.execute("SELECT DISTINCT 姓名, 月份 FROM salary_table WHERE 月份 != ''")
            salary_records = cursor.fetchall()

            synced_count = 0
            error_count = 0
            errors = []

            for record in salary_records:
                name = record['姓名']
                month = record['月份'] or ''

                try:
                    cursor.execute(
                        "SELECT 手机号, 部门, 入职日期 FROM users WHERE 姓名 = ?",
                        (name,)
                    )
                    user_info = cursor.fetchone()
                    phone_number = user_info['手机号'] if user_info and user_info['手机号'] else None
                    user_dept = user_info['部门'] if user_info and user_info['部门'] else None
                    user_hire_date = user_info['入职日期'] if user_info and user_info['入职日期'] else None

                    cursor.execute("SELECT * FROM salary_table WHERE 姓名 = ? AND 月份 = ?", (name, month))
                    salary_data = cursor.fetchone()

                    if not salary_data:
                        errors.append(f"用户 {name} ({month}) 的工资数据不存在")
                        error_count += 1
                        continue

                    cursor.execute(
                        "SELECT id FROM summary_table WHERE 姓名 = ? AND 月份 = ?",
                        (name, month)
                    )
                    existing = cursor.fetchone()

                    if existing:
                        cursor.execute(
                            """UPDATE summary_table
                               SET 序号 = ?,
                                   部门 = COALESCE(?, 部门),
                                   入职日期 = COALESCE(?, 入职日期),
                                   手机号 = COALESCE(?, 手机号),
                                   是否代扣社保 = ?,
                                   岗位 = ?,
                                   应出勤天数 = ?,
                                   实际出勤天数 = ?,
                                   上门服务小时 = ?,
                                   基本工资底薪 = ?,
                                   基本工资其它补贴 = ?,
                                   基本工资合计 = ?,
                                   岗位工资 = ?,
                                   交通费 = ?,
                                   手机费 = ?,
                                   奖金 = ?,
                                   高温费 = ?,
                                   护理员绩效工资 = ?,
                                   应发工资 = ?,
                                   应扣款项缺勤扣款 = ?,
                                   应扣款项养老 = ?,
                                   应扣款项医疗 = ?,
                                   应扣款项失业 = ?,
                                   应扣款项公积金 = ?,
                                   应扣款项应缴个税 = ?,
                                   其它扣款 = ?,
                                   住宿扣款 = ?,
                                   水电扣款 = ?,
                                   实发工资 = ?,
                                   月份 = ?,
                                   updated_at = ?
                               WHERE 姓名 = ? AND 月份 = ?""",
                            (
                                salary_data.get('序号'),
                                user_dept or salary_data.get('部门'),
                                user_hire_date or salary_data.get('入职日期'),
                                phone_number,
                                salary_data.get('是否代扣社保'),
                                salary_data.get('岗位'),
                                salary_data.get('应出勤天数'),
                                salary_data.get('实际出勤天数'),
                                salary_data.get('上门服务小时'),
                                salary_data.get('基本工资底薪'),
                                salary_data.get('基本工资其它补贴'),
                                salary_data.get('基本工资合计'),
                                salary_data.get('岗位工资'),
                                salary_data.get('交通费'),
                                salary_data.get('手机费'),
                                salary_data.get('奖金'),
                                salary_data.get('高温费'),
                                salary_data.get('护理员绩效工资'),
                                salary_data.get('应发工资'),
                                salary_data.get('应扣款项缺勤扣款'),
                                salary_data.get('应扣款项养老'),
                                salary_data.get('应扣款项医疗'),
                                salary_data.get('应扣款项失业'),
                                salary_data.get('应扣款项公积金'),
                                salary_data.get('应扣款项应缴个税'),
                                salary_data.get('其它扣款'),
                                salary_data.get('住宿扣款'),
                                salary_data.get('水电扣款'),
                                salary_data.get('实发工资'),
                                month,
                                datetime.now(),
                                name, month
                            )
                        )
                    else:
                        cursor.execute(
                            """INSERT INTO summary_table
                               (序号, 部门, 姓名, 入职日期, 手机号, 是否代扣社保, 岗位,
                                应出勤天数, 实际出勤天数, 上门服务小时,
                                基本工资底薪, 基本工资其它补贴, 基本工资合计,
                                岗位工资, 交通费, 手机费, 奖金, 高温费,
                                护理员绩效工资, 应发工资,
                                应扣款项缺勤扣款, 应扣款项养老, 应扣款项医疗,
                                应扣款项失业, 应扣款项公积金, 应扣款项应缴个税,
                                其它扣款, 住宿扣款, 水电扣款,
                                实发工资, 月份, 查询状态, 查询次数)
                               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '已查询', 1)""",
                            (
                                salary_data.get('序号'),
                                user_dept or salary_data.get('部门'),
                                name,
                                user_hire_date or salary_data.get('入职日期'),
                                phone_number,
                                salary_data.get('是否代扣社保'),
                                salary_data.get('岗位'),
                                salary_data.get('应出勤天数'),
                                salary_data.get('实际出勤天数'),
                                salary_data.get('上门服务小时'),
                                salary_data.get('基本工资底薪'),
                                salary_data.get('基本工资其它补贴'),
                                salary_data.get('基本工资合计'),
                                salary_data.get('岗位工资'),
                                salary_data.get('交通费'),
                                salary_data.get('手机费'),
                                salary_data.get('奖金'),
                                salary_data.get('高温费'),
                                salary_data.get('护理员绩效工资'),
                                salary_data.get('应发工资'),
                                salary_data.get('应扣款项缺勤扣款'),
                                salary_data.get('应扣款项养老'),
                                salary_data.get('应扣款项医疗'),
                                salary_data.get('应扣款项失业'),
                                salary_data.get('应扣款项公积金'),
                                salary_data.get('应扣款项应缴个税'),
                                salary_data.get('其它扣款'),
                                salary_data.get('住宿扣款'),
                                salary_data.get('水电扣款'),
                                salary_data.get('实发工资'),
                                month
                            )
                        )

                    synced_count += 1

                except Exception as e:
                    errors.append(f"同步用户 {name} ({month}) 失败: {str(e)}")
                    error_count += 1

            conn.commit()

            print(f"[INFO] 数据汇总完成: 成功 {synced_count} 条, 失败 {error_count} 条")

            return {
                'success': True,
                'synced_count': synced_count,
                'error_count': error_count,
                'errors': errors
            }

        except Exception as e:
            print(f"[ERROR] 数据汇总失败: {e}")
            return {
                'success': False,
                'synced_count': 0,
                'error_count': 0,
                'errors': [str(e)]
            }
        finally:
            if conn:
                conn.close()

    def import_salary_from_csv(self, csv_file_path: str, month: str = '', has_header: bool = True) -> Dict:
        """
        从CSV文件导入工资数据
        
        Args:
            csv_file_path: CSV文件路径
            has_header: 是否包含表头
        
        Returns:
            dict: {
                'success': bool,
                'imported_count': int,
                'error_count': int,
                'errors': list
            }
        """
        conn = None
        try:
            if not os.path.exists(csv_file_path):
                return {
                    'success': False,
                    'imported_count': 0,
                    'error_count': 0,
                    'errors': ['文件不存在']
                }
            
            conn = self.get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute("SELECT 姓名 FROM users")
            existing_users = set(row['姓名'] for row in cursor.fetchall())
            
            new_users = set()
            
            imported_count = 0
            error_count = 0
            errors = []
            
            with open(csv_file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                
                # 读取表头
                headers = []
                if has_header:
                    headers = next(reader)
                    headers = [h.strip() for h in headers]
                else:
                    errors.append("CSV文件必须包含表头")
                    return {
                        'success': False,
                        'imported_count': 0,
                        'error_count': 1,
                        'errors': errors
                    }
                
                # 检查必要字段
                if '姓名' not in headers:
                    errors.append("CSV文件缺少'姓名'列")
                    return {
                        'success': False,
                        'imported_count': 0,
                        'error_count': 1,
                        'errors': errors
                    }
                
                # 获取姓名列索引
                name_col_idx = headers.index('姓名')
                
                for row_num, row in enumerate(reader, start=2):
                    try:
                        # 获取姓名
                        name = row[name_col_idx].strip() if name_col_idx < len(row) else ''
                        
                        if not name:
                            errors.append(f"第{row_num}行: 姓名为空")
                            error_count += 1
                            continue
                        
                        if name not in existing_users:
                            new_users.add(name)
                        
                        # 构建数据字典
                        row_data = {}
                        for idx, header in enumerate(headers):
                            if idx < len(row):
                                value = row[idx].strip()
                                if value:
                                    row_data[header] = value
                        
                        # 检查是否已存在
                        cursor.execute(
                            "SELECT id FROM salary_table WHERE 姓名 = ? AND 月份 = ?",
                            (name, month)
                        )
                        
                        existing = cursor.fetchone()
                        
                        # 定义所有可能的字段
                        all_fields = ['序号', '部门', '入职日期', '应出勤天数', '实际出勤天数', 
                                     '上门服务小时', '基本工资底薪', '基本工资绩效', '基本工资合计',
                                     '岗位工资', '护理员绩效工资', '应发工资', '应扣款项社保',
                                     '应扣款项公积金', '应扣款项个税', '应扣款项其他',
                                     '应扣款项合计', '实发工资', '是否代扣社保']
                        
                        if existing:
                            # 更新现有记录
                            update_fields = []
                            update_values = []
                            
                            for field in all_fields:
                                if field in row_data and field != '姓名':
                                    update_fields.append(f"{field} = ?")
                                    # 根据字段类型处理值
                                    if field in ['应出勤天数', '实际出勤天数', '上门服务小时', 
                                                '基本工资底薪', '基本工资绩效', '基本工资合计',
                                                '岗位工资', '护理员绩效工资', '应发工资', 
                                                '应扣款项社保', '应扣款项公积金', '应扣款项个税',
                                                '应扣款项其他', '应扣款项合计', '实发工资']:
                                        update_values.append(self._parse_float(row_data[field]))
                                    else:
                                        update_values.append(row_data[field])
                            
                            if update_fields:
                                update_fields.append("updated_at = ?")
                                update_values.append(datetime.now())
                                update_values.append(name)
                                update_values.append(month)
                                
                                sql = f"UPDATE salary_table SET {', '.join(update_fields)} WHERE 姓名 = ? AND 月份 = ?"
                                cursor.execute(sql, update_values)
                        else:
                            fields = ['姓名', '月份']
                            placeholders = ['?', '?']
                            values = [name, month]
                            
                            for field in all_fields:
                                if field in row_data:
                                    fields.append(field)
                                    placeholders.append('?')
                                    # 根据字段类型处理值
                                    if field in ['应出勤天数', '实际出勤天数', '上门服务小时', 
                                                '基本工资底薪', '基本工资绩效', '基本工资合计',
                                                '岗位工资', '护理员绩效工资', '应发工资', 
                                                '应扣款项社保', '应扣款项公积金', '应扣款项个税',
                                                '应扣款项其他', '应扣款项合计', '实发工资']:
                                        values.append(self._parse_float(row_data[field]))
                                    else:
                                        values.append(row_data[field])
                            
                            sql = f"INSERT INTO salary_table ({', '.join(fields)}) VALUES ({', '.join(placeholders)})"
                            cursor.execute(sql, values)
                        
                        imported_count += 1
                        
                    except Exception as e:
                        errors.append(f"第{row_num}行: {str(e)}")
                        error_count += 1
            
            conn.commit()
            
            print(f"[INFO] CSV导入完成: 成功 {imported_count} 条, 失败 {error_count} 条")
            
            result = {
                'success': True,
                'imported_count': imported_count,
                'error_count': error_count,
                'errors': errors
            }
            
            if new_users:
                result['new_users'] = list(new_users)
                result['message'] = f'导入成功 {imported_count} 条记录。发现 {len(new_users)} 个新用户：{", ".join(sorted(new_users))}。这些用户的工资信息暂不显示，请前往"人员管理"页面创建账号后即可显示。'
            
            return result
            
        except Exception as e:
            print(f"[ERROR] CSV导入失败: {e}")
            return {
                'success': False,
                'imported_count': 0,
                'error_count': 0,
                'errors': [str(e)]
            }
        finally:
            if conn:
                conn.close()
    
    def export_summary_to_csv(self, output_path: str, department: str = None) -> Dict:
        """
        导出汇总数据到CSV文件
        
        Args:
            output_path: 输出文件路径
            department: 部门名称（可选，不指定则导出全部）
        
        Returns:
            dict: {
                'success': bool,
                'exported_count': int,
                'file_path': str
            }
        """
        conn = None
        try:
            conn = self.get_db_connection()
            cursor = conn.cursor()
            
            # 查询汇总数据
            if department:
                cursor.execute(
                    """SELECT * FROM summary_table WHERE 部门 = ? ORDER BY 序号""",
                    (department,)
                )
            else:
                cursor.execute("SELECT * FROM summary_table ORDER BY 部门, 序号")
            
            summaries = cursor.fetchall()
            
            # 写入CSV文件
            with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                
                # 写入表头
                writer.writerow([
                    '序号', '部门', '姓名', '入职日期', '手机号', '是否代扣社保', '岗位',
                    '应出勤（天）', '实际出勤（天）', '上门服务（小时）',
                    '基本工资-底薪', '基本工资-其它补贴', '基本工资-基本工资',
                    '岗位工资', '交通费', '手机费', '奖金', '高温费', '护理员绩效工资',
                    '应发工资',
                    '应扣款项-缺勤扣款', '应扣款项-养老(8%)', '应扣款项-医疗(2%)',
                    '应扣款项-失业(0.5%)', '应扣款项-公积金(7%)', '应扣款项-应缴个税',
                    '其它扣款', '住宿扣款', '水电扣款',
                    '实发工资',
                    '月份', '查询状态', '查询次数', '签收状态', '最新签收时间', '签收方式', '签名图片'
                ])
                
                # 写入数据
                for summary in summaries:
                    writer.writerow([
                        summary['序号'],
                        summary['部门'],
                        summary['姓名'],
                        str(summary['入职日期'])[:10] if summary['入职日期'] else '',
                        summary['手机号'],
                        summary['是否代扣社保'],
                        summary['岗位'],
                        summary['应出勤天数'],
                        summary['实际出勤天数'],
                        summary['上门服务小时'],
                        summary['基本工资底薪'],
                        summary['基本工资其它补贴'],
                        summary['基本工资合计'],
                        summary['岗位工资'],
                        summary['交通费'],
                        summary['手机费'],
                        summary['奖金'],
                        summary['高温费'],
                        summary['护理员绩效工资'],
                        summary['应发工资'],
                        summary['应扣款项缺勤扣款'],
                        summary['应扣款项养老'],
                        summary['应扣款项医疗'],
                        summary['应扣款项失业'],
                        summary['应扣款项公积金'],
                        summary['应扣款项应缴个税'],
                        summary['其它扣款'],
                        summary['住宿扣款'],
                        summary['水电扣款'],
                        summary['实发工资'],
                        summary['月份'],
                        summary['查询状态'],
                        summary['查询次数'],
                        summary['签收状态'],
                        summary['最新签收时间'],
                        summary['签收方式'],
                        summary['签名图片']
                    ])
            
            print(f"[INFO] CSV导出完成: {len(summaries)} 条数据")
            
            return {
                'success': True,
                'exported_count': len(summaries),
                'file_path': output_path
            }
            
        except Exception as e:
            print(f"[ERROR] CSV导出失败: {e}")
            return {
                'success': False,
                'exported_count': 0,
                'file_path': ''
            }
        finally:
            if conn:
                conn.close()
    
    def export_summary_to_excel(self, output_path: str, department: str = None, month: str = None) -> Dict:
        conn = None
        try:
            import openpyxl
            from openpyxl.drawing.image import Image as XlImage
            from openpyxl.utils import get_column_letter
            
            conn = self.get_db_connection()
            cursor = conn.cursor()
            
            conditions = []
            params = []
            
            if department:
                conditions.append("s.部门 = ?")
                params.append(department)
            
            if month:
                conditions.append("s.月份 = ?")
                params.append(month)
            
            where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""
            
            cursor.execute(
                f"""SELECT 
                    s.序号, s.部门, s.姓名, s.入职日期, u.手机号, s.是否代扣社保, s.岗位,
                    s.应出勤天数, s.实际出勤天数, s.上门服务小时,
                    s.基本工资底薪, s.基本工资其它补贴, s.基本工资合计,
                    s.岗位工资, s.交通费, s.手机费, s.奖金, s.高温费, s.护理员绩效工资,
                    s.应发工资,
                    s.应扣款项缺勤扣款, s.应扣款项养老, s.应扣款项医疗,
                    s.应扣款项失业, s.应扣款项公积金, s.应扣款项应缴个税,
                    s.其它扣款, s.住宿扣款, s.水电扣款,
                    s.实发工资,
                    s.月份,
                    st.查询状态, st.查询次数, st.签收状态, st.最新签收时间, st.签收方式, st.签名图片
                FROM salary_table s
                LEFT JOIN users u ON s.姓名 = u.姓名
                LEFT JOIN summary_table st ON s.姓名 = st.姓名 AND s.月份 = st.月份
                {where_clause}
                ORDER BY s.月份 DESC, s.部门, CAST(s.序号 AS INTEGER)""",
                params
            )
            
            summaries = cursor.fetchall()
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "工资汇总表"
            
            headers = [
                '序号', '部门', '姓名', '入职日期', '手机号', '是否代扣社保', '岗位',
                '应出勤（天）', '实际出勤（天）', '上门服务（小时）',
                '基本工资-底薪', '基本工资-其它补贴', '基本工资-基本工资',
                '岗位工资', '交通费', '手机费', '奖金', '高温费', '护理员绩效工资',
                '应发工资',
                '应扣款项-缺勤扣款', '应扣款项-养老(8%)', '应扣款项-医疗(2%)',
                '应扣款项-失业(0.5%)', '应扣款项-公积金(7%)', '应扣款项-应缴个税',
                '其它扣款', '住宿扣款', '水电扣款',
                '实发工资',
                '月份', '查询状态', '查询次数', '签收状态', '最新签收时间', '签收方式', '签名图片'
            ]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            signature_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'backend', 'signatures')
            if not os.path.exists(signature_dir):
                signature_dir = os.path.join(os.path.dirname(__file__), 'signatures')
            
            for row_idx, summary in enumerate(summaries, 2):
                hire_date = summary['入职日期']
                hire_date_str = str(hire_date)[:10] if hire_date else ''
                
                latest_sign_time = summary['最新签收时间']
                sign_time_str = str(latest_sign_time)[:19] if latest_sign_time else ''
                
                row_data = [
                    summary['序号'],
                    summary['部门'],
                    summary['姓名'],
                    hire_date_str,
                    summary['手机号'],
                    summary['是否代扣社保'],
                    summary['岗位'],
                    summary['应出勤天数'],
                    summary['实际出勤天数'],
                    summary['上门服务小时'],
                    summary['基本工资底薪'],
                    summary['基本工资其它补贴'],
                    summary['基本工资合计'],
                    summary['岗位工资'],
                    summary['交通费'],
                    summary['手机费'],
                    summary['奖金'],
                    summary['高温费'],
                    summary['护理员绩效工资'],
                    summary['应发工资'],
                    summary['应扣款项缺勤扣款'],
                    summary['应扣款项养老'],
                    summary['应扣款项医疗'],
                    summary['应扣款项失业'],
                    summary['应扣款项公积金'],
                    summary['应扣款项应缴个税'],
                    summary['其它扣款'],
                    summary['住宿扣款'],
                    summary['水电扣款'],
                    summary['实发工资'],
                    summary['月份'],
                    summary['查询状态'] or '未查询',
                    summary['查询次数'] or 0,
                    summary['签收状态'] or '未签收',
                    sign_time_str,
                    summary['签收方式'],
                    ''
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                
                signature_name = summary['签名图片']
                if signature_name:
                    print(f"[DEBUG] 查找签名: signature_name={signature_name}, signature_dir={signature_dir}, dir_exists={os.path.exists(signature_dir)}")
                    sig_file = None
                    if os.path.exists(signature_dir):
                        all_files = os.listdir(signature_dir)
                        print(f"[DEBUG] 签名目录文件: {all_files[:10]}...")
                        for f in all_files:
                            if f.startswith(signature_name + '.') and f.lower().endswith(('.png', '.jpg', '.jpeg')):
                                sig_file = f
                                print(f"[DEBUG] 找到签名文件: {sig_file}")
                                break
                    
                    if sig_file:
                        img_path = os.path.join(signature_dir, sig_file)
                        print(f"[DEBUG] 图片路径: {img_path}, exists={os.path.exists(img_path)}")
                        try:
                            img = XlImage(img_path)
                            img.width = 80
                            img.height = 40
                            sig_col = len(headers)
                            cell_ref = f"{get_column_letter(sig_col)}{row_idx}"
                            ws.row_dimensions[row_idx].height = 35
                            ws.add_image(img, cell_ref)
                            print(f"[DEBUG] 图片插入成功: {cell_ref}")
                        except Exception as e:
                            print(f"[WARN] 插入签名图片失败 {img_path}: {e}")
                            import traceback
                            traceback.print_exc()
                            ws.cell(row=row_idx, column=len(headers), value='已签名')
                    else:
                        print(f"[DEBUG] 未找到签名文件: {signature_name}")
                        ws.cell(row=row_idx, column=len(headers), value='已签名')
            
            for col_idx in range(1, len(headers) + 1):
                max_length = max(
                    len(str(ws.cell(row=r, column=col_idx).value or ''))
                    for r in range(1, min(len(summaries) + 2, 20))
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 25)
            
            sig_col_letter = get_column_letter(len(headers))
            ws.column_dimensions[sig_col_letter].width = 15
            
            wb.save(output_path)
            
            print(f"[INFO] Excel导出完成: {len(summaries)} 条数据")
            
            return {
                'success': True,
                'exported_count': len(summaries),
                'file_path': output_path
            }
            
        except Exception as e:
            print(f"[ERROR] Excel导出失败: {e}")
            return {
                'success': False,
                'exported_count': 0,
                'file_path': ''
            }
        finally:
            if conn:
                conn.close()
    
    def export_statistics_report(self, output_path: str) -> Dict:
        """
        导出统计报表
        
        Args:
            output_path: 输出文件路径
        
        Returns:
            dict: {
                'success': bool,
                'file_path': str
            }
        """
        conn = None
        try:
            conn = self.get_db_connection()
            cursor = conn.cursor()
            
            # 查询总体统计
            cursor.execute(
                """SELECT 
                    COUNT(*) as total_count,
                    SUM(CASE WHEN 查询状态 = '已查询' THEN 1 ELSE 0 END) as queried_count,
                    SUM(CASE WHEN 签收状态 = '已签收' THEN 1 ELSE 0 END) as signed_count,
                    SUM(查询次数) as total_queries,
                    AVG(实发工资) as avg_salary,
                    SUM(实发工资) as total_salary
                   FROM summary_table"""
            )
            
            overall_stats = cursor.fetchone()
            
            # 查询部门统计
            cursor.execute(
                """SELECT 
                    部门,
                    COUNT(*) as total_count,
                    SUM(CASE WHEN 查询状态 = '已查询' THEN 1 ELSE 0 END) as queried_count,
                    SUM(CASE WHEN 签收状态 = '已签收' THEN 1 ELSE 0 END) as signed_count,
                    SUM(查询次数) as total_queries,
                    AVG(实发工资) as avg_salary,
                    SUM(实发工资) as total_salary
                   FROM summary_table
                   GROUP BY 部门
                   ORDER BY 部门"""
            )
            
            department_stats = cursor.fetchall()
            
            # 写入报表
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write("=" * 80 + "\n")
                f.write("工资系统统计报表\n")
                f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("=" * 80 + "\n\n")
                
                # 总体统计
                f.write("一、总体统计\n")
                f.write("-" * 80 + "\n")
                f.write(f"总人数: {overall_stats['total_count']}\n")
                f.write(f"已查询人数: {overall_stats['queried_count']}\n")
                f.write(f"已签收人数: {overall_stats['signed_count']}\n")
                f.write(f"总查询次数: {overall_stats['total_queries'] or 0}\n")
                f.write(f"查询率: {(overall_stats['queried_count'] / overall_stats['total_count'] * 100) if overall_stats['total_count'] > 0 else 0:.2f}%\n")
                f.write(f"签收率: {(overall_stats['signed_count'] / overall_stats['total_count'] * 100) if overall_stats['total_count'] > 0 else 0:.2f}%\n")
                f.write(f"平均工资: {overall_stats['avg_salary'] or 0:.2f}元\n")
                f.write(f"工资总额: {overall_stats['total_salary'] or 0:.2f}元\n")
                f.write("\n")
                
                # 部门统计
                f.write("二、部门统计\n")
                f.write("-" * 80 + "\n")
                
                for dept in department_stats:
                    f.write(f"\n部门: {dept['部门']}\n")
                    f.write(f"  总人数: {dept['total_count']}\n")
                    f.write(f"  已查询: {dept['queried_count']}\n")
                    f.write(f"  已签收: {dept['signed_count']}\n")
                    f.write(f"  总查询次数: {dept['total_queries'] or 0}\n")
                    query_rate = (dept['queried_count'] / dept['total_count'] * 100) if dept['total_count'] > 0 else 0
                    sign_rate = (dept['signed_count'] / dept['total_count'] * 100) if dept['total_count'] > 0 else 0
                    f.write(f"  查询率: {query_rate:.2f}%\n")
                    f.write(f"  签收率: {sign_rate:.2f}%\n")
                    f.write(f"  平均工资: {dept['avg_salary'] or 0:.2f}元\n")
                    f.write(f"  工资总额: {dept['total_salary'] or 0:.2f}元\n")
                
                f.write("\n" + "=" * 80 + "\n")
            
            print(f"[INFO] 统计报表导出完成: {output_path}")
            
            return {
                'success': True,
                'file_path': output_path
            }
            
        except Exception as e:
            print(f"[ERROR] 统计报表导出失败: {e}")
            return {
                'success': False,
                'file_path': ''
            }
        finally:
            if conn:
                conn.close()
    
    def _parse_float(self, value: str) -> float:
        """解析浮点数"""
        try:
            return float(value.strip()) if value.strip() else 0.0
        except:
            return 0.0
    
    def update_salary_data(self, name: str, updates: Dict) -> bool:
        """
        更新工资数据
        
        Args:
            name: 用户姓名
            updates: 更新字段字典
        
        Returns:
            bool: 是否成功
        """
        conn = None
        try:
            conn = self.get_db_connection()
            cursor = conn.cursor()
            
            # 构建更新SQL
            set_clauses = []
            values = []
            
            for field, value in updates.items():
                set_clauses.append(f"{field} = ?")
                values.append(value)
            
            set_clauses.append("updated_at = ?")
            values.append(datetime.now())
            values.append(name)
            
            sql = f"UPDATE salary_table SET {', '.join(set_clauses)} WHERE 姓名 = ?"
            
            cursor.execute(sql, values)
            
            conn.commit()
            
            print(f"[INFO] 工资数据更新成功: {name}")
            return True
            
        except Exception as e:
            print(f"[ERROR] 更新工资数据失败: {e}")
            return False
        finally:
            if conn:
                conn.close()
    
    def get_data_summary(self) -> Dict:
        """
        获取数据汇总信息
        
        Returns:
            dict: 数据汇总信息
        """
        conn = None
        try:
            conn = self.get_db_connection()
            cursor = conn.cursor()
            
            # 工资表统计
            cursor.execute("SELECT COUNT(*) FROM salary_table")
            salary_count = cursor.fetchone()[0]
            
            # 汇总表统计
            cursor.execute("SELECT COUNT(*) FROM summary_table")
            summary_count = cursor.fetchone()[0]
            
            # 用户表统计
            cursor.execute("SELECT COUNT(*) FROM users WHERE role = 'user'")
            user_count = cursor.fetchone()[0]
            
            # 签名文件统计
            cursor.execute("SELECT COUNT(*) FROM signature_files WHERE 文件状态 = '正常'")
            signature_count = cursor.fetchone()[0]
            
            # 查询统计
            cursor.execute(
                """SELECT 
                    SUM(CASE WHEN 查询状态 = '已查询' THEN 1 ELSE 0 END) as queried_count,
                    SUM(CASE WHEN 签收状态 = '已签收' THEN 1 ELSE 0 END) as signed_count,
                    SUM(查询次数) as total_queries
                   FROM summary_table"""
            )
            
            query_stats = cursor.fetchone()
            
            return {
                'salary_count': salary_count,
                'summary_count': summary_count,
                'user_count': user_count,
                'signature_count': signature_count,
                'queried_count': query_stats['queried_count'] or 0,
                'signed_count': query_stats['signed_count'] or 0,
                'total_queries': query_stats['total_queries'] or 0
            }
            
        except Exception as e:
            print(f"[ERROR] 获取数据汇总信息失败: {e}")
            return {
                'salary_count': 0,
                'summary_count': 0,
                'user_count': 0,
                'signature_count': 0,
                'queried_count': 0,
                'signed_count': 0,
                'total_queries': 0
            }
        finally:
            if conn:
                conn.close()
