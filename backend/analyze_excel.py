#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
分析Excel文件结构
"""

import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_excel(file_path):
    print(f"\n{'='*60}")
    print(f"分析文件: {file_path}")
    print(f"{'='*60}\n")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        print(f"工作表名称: {ws.title}")
        print(f"总行数: {ws.max_row}")
        print(f"总列数: {ws.max_column}")
        
        print(f"\n{'='*60}")
        print("前5行内容:")
        print(f"{'='*60}\n")
        
        for row_num in range(1, min(6, ws.max_row + 1)):
            row_values = []
            for col_num in range(1, min(30, ws.max_column + 1)):
                cell = ws.cell(row=row_num, column=col_num)
                value = cell.value
                row_values.append(str(value) if value is not None else '(空)')
            print(f"第{row_num}行: {row_values}")
        
        print(f"\n{'='*60}")
        print("表头分析 (第1行):")
        print(f"{'='*60}\n")
        
        headers = []
        for col_num in range(1, min(30, ws.max_column + 1)):
            cell = ws.cell(row=1, column=col_num)
            value = cell.value
            headers.append(str(value) if value is not None else f'列{col_num}')
            print(f"列{col_num}: '{value}'")
        
        print(f"\n{'='*60}")
        print("检查必要字段:")
        print(f"{'='*60}\n")
        
        required_fields = ['姓名', '入职日期', '部门', '序号']
        for field in required_fields:
            found = False
            for idx, h in enumerate(headers):
                if field in h.replace(' ', '').replace('（', '(').replace('）', ')'):
                    print(f"[OK] '{field}' 匹配到 '{h}' (列 {idx + 1})")
                    found = True
                    break
            if not found:
                print(f"[X] '{field}' 不存在")
        
        print(f"\n{'='*60}")
        print("数据行分析 (第2-5行):")
        print(f"{'='*60}\n")
        
        for row_num in range(2, min(6, ws.max_row + 1)):
            print(f"\n--- 第{row_num}行 ---")
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                value = cell.value
                if value is not None:
                    print(f"  {header}: {value}")
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    test_files = [
        r'C:\Users\yao\Desktop\工资签收系统\工资表.xlsx',
    ]
    
    for file_path in test_files:
        try:
            analyze_excel(file_path)
        except FileNotFoundError:
            print(f"文件不存在: {file_path}")
        print("\n")
